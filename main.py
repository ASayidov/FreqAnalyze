import os
import sys
import re
import pandas as pd
from tkinter import Tk, Button, Label, Entry, messagebox, Toplevel, filedialog
from geopy.distance import geodesic
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Файл йўлини тўғри созлайди, агар .exe бўлса
def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# TRXNAMEга асосланган базалар, секторлар ва передатчикларни ажратиш функцияси
def extract_base_sector_transmitter(trxname, region_code):
    trxname_cleaned = trxname[len(region_code):] if trxname.startswith(region_code) else trxname

    if trxname_cleaned[0] == "0":
        trxname_cleaned = trxname_cleaned[1:]

    if len(trxname_cleaned) == 6:
        base_number = trxname_cleaned[:4]
        sector = trxname_cleaned[4]
        transmitter = trxname_cleaned[5]
    elif len(trxname_cleaned) == 5:
        base_number = trxname_cleaned[:3]
        sector = trxname_cleaned[3]
        transmitter = trxname_cleaned[4]
    else:
        raise ValueError(f"Invalid TRXNAME format: {trxname}.")

    return base_number, sector, transmitter

# testsa.py учун функциялар
def parse_text_file_testsa(file_path, region_code):
    results = []
    with open(file_path, 'r') as file:
        lines = file.readlines()
        for line in lines:
            if line.startswith("ADD GTRX"):
                trxname_match = re.search(r'TRXNAME="(\d+)"', line)
                freq_match = re.search(r'FREQ=(\d+)', line)
                ismainbcch_match = re.search(r'ISMAINBCCH=(YES|NO)', line)

                if trxname_match and freq_match and ismainbcch_match:
                    trxname = trxname_match.group(1)
                    freq = int(freq_match.group(1))
                    ismainbcch = ismainbcch_match.group(1) == "YES"

                    try:
                        base_number, sector, transmitter = extract_base_sector_transmitter(trxname, region_code)
                    except ValueError as e:
                        print(e)
                        continue

                    freq_type = "BCCH" if ismainbcch else "TCH"

                    results.append({
                        "Base Number": base_number,
                        "Sector": sector,
                        "Transmitter": transmitter,
                        "Frequency": freq,
                        "Frequency Type": freq_type
                    })
    return results

# testne.py учун функциялар
def parse_text_file_testne(file_path, region_code):
    results = []
    with open(file_path, 'r') as file:
        lines = file.readlines()
        for line in lines:
            if line.startswith("ADD GTRX"):
                trxname_match = re.search(r'TRXNAME="(\d+)"', line)
                freq_match = re.search(r'FREQ=(\d+)', line)
                ismainbcch_match = re.search(r'ISMAINBCCH=(YES|NO)', line)

                if trxname_match and freq_match and ismainbcch_match:
                    trxname = trxname_match.group(1)
                    freq = int(freq_match.group(1))
                    ismainbcch = ismainbcch_match.group(1) == "YES"

                    try:
                        base_number, sector, transmitter = extract_base_sector_transmitter(trxname, region_code)
                    except ValueError as e:
                        print(e)
                        continue

                    freq_type = "BCCH" if ismainbcch else "TCH"

                    results.append({
                        "Base Number": base_number,
                        "Sector": sector,
                        "Transmitter": transmitter,
                        "Frequency": freq,
                        "Frequency Type": freq_type
                    })
    return results

# "О программе" ойнаси
def show_about():
    about_text = (
        "Программа: Анализатор Частот GSM\n"
        "Версия: 1.0\n"
        "Цель программы: Анализ частот для GSM-сетей,\n"
        "включая поиск одинаковых и смежных частот в заданном радиусе.\n\n"
        "Автор: Абдуллажон Сайидов\n"
        "Электронная почта: abdullajonsayidov@gmail.com\n"
        "Технические требования: Windows 10/11."
    )
    messagebox.showinfo("О программе", about_text)

def open_testsa_window():
    new_window = Toplevel(root)
    new_window.title("Testsa Analysis")
    new_window.transient(root)  # Асосий ойна орқада қолиши учун
    FrequencyAnalyzerApp(new_window, parse_text_file_function=parse_text_file_testsa, analyze_function=analyze_same_frequencies, title="Определение одинаковых частот")

def open_testne_window():
    new_window = Toplevel(root)
    new_window.title("Testne Analysis")
    new_window.transient(root)  # Асосий ойна орқада қолиши учун
    FrequencyAnalyzerApp(new_window, parse_text_file_function=parse_text_file_testne, analyze_function=analyze_adjacent_frequencies, title="Определение смежных частот")

class FrequencyAnalyzerApp:
    def __init__(self, root, parse_text_file_function=None, analyze_function=None, title=""):
        self.root = root
        self.parse_text_file_function = parse_text_file_function
        self.analyze_function = analyze_function
        self.root.title(title)

        self.root.geometry("600x700")  # 600 пиксел кенглик, 700 пиксел баландлик

        self.title_label = Label(self.root, text=title, font=("Arial", 16))
        self.title_label.pack(pady=10)

        self.region_code_label = Label(self.root, text="Введите код региона (например, 1515)")
        self.region_code_label.pack(pady=5)

        self.region_code_input = Entry(self.root)
        self.region_code_input.pack(pady=5)

        self.load_text_button1 = Button(self.root, text="Выбрать MMLCFG файл (Text 1)", command=lambda: self.load_text_file(1))
        self.load_text_button1.pack(pady=10)

        self.load_text_button2 = Button(self.root, text="Выбрать MMLCFG файл (Text 2)", command=lambda: self.load_text_file(2))
        self.load_text_button2.pack(pady=10)

        self.load_text_button3 = Button(self.root, text="Выбрать MMLCFG файл (Text 3)", command=lambda: self.load_text_file(3))
        self.load_text_button3.pack(pady=10)

        self.load_text_button4 = Button(self.root, text="Выбрать MMLCFG файл (Text 4)", command=lambda: self.load_text_file(4))
        self.load_text_button4.pack(pady=10)

        self.load_excel_button = Button(self.root, text="Выбрать forKMZ файл (Excel)", command=self.load_excel_file)
        self.load_excel_button.pack(pady=10)

        self.distance_label = Label(self.root, text="Введите расстояние в метрах")
        self.distance_label.pack(pady=5)

        self.distance_input = Entry(self.root)
        self.distance_input.pack(pady=5)

        self.analyze_button = Button(self.root, text="Анализировать", command=self.analyze_data)
        self.analyze_button.pack(pady=20)

        self.save_button = Button(self.root, text="Сохранить результаты", command=self.save_results)
        self.save_button.pack(pady=10)

        self.file_path_label = Label(self.root, text="Путь к файлу: Не выбран")
        self.file_path_label.pack(pady=10)

        self.status_label = Label(self.root, text="Статус: Ожидание ввода.")
        self.status_label.pack(pady=10)

        self.text_data = []
        self.excel_data = None
        self.analysis_results = None

    def load_text_file(self, index):
        file_path = filedialog.askopenfilename(title="Открыть текстовый файл", filetypes=(("Text Files", "*.txt"),))
        if file_path:
            region_code = self.region_code_input.get()
            parsed_data = self.parse_text_file_function(file_path, region_code)
            self.text_data.extend(parsed_data)
            self.file_path_label.config(text=f"Text File {index}: {file_path}")
            self.status_label.config(text=f"Text file {index} loaded successfully!")

    def load_excel_file(self):
        file_path = filedialog.askopenfilename(title="Открыть файл Excel", filetypes=(("Excel Files", "*.xlsx"),))
        if file_path:
            self.excel_data = process_excel_file(file_path)
            self.file_path_label.config(text=f"Excel File: {file_path}")
            self.status_label.config(text="Excel file loaded successfully!")

    def analyze_data(self):
        if self.text_data and self.excel_data:
            try:
                distance_limit = float(self.distance_input.get())
                if distance_limit <= 0:
                    raise ValueError("Distance must be a positive number.")
            except ValueError as e:
                self.status_label.config(text=f"Invalid distance value: {str(e)}")
                return

            self.analysis_results = self.analyze_function(self.text_data, self.excel_data, distance_limit, self.status_label)
            if self.analysis_results:
                self.status_label.config(text="Analysis complete!")
            else:
                self.status_label.config(text="No results found in analysis.")
        else:
            self.status_label.config(text="Please load at least one text file and the Excel file first.")

    def save_results(self):
        if self.analysis_results:
            output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))

            if output_path:
                try:
                    save_results(self.analysis_results, output_path)
                    self.file_path_label.config(text=f"Saved File: {output_path}")
                    self.status_label.config(text="Results saved successfully!")
                except Exception as e:
                    self.status_label.config(text=f"Error saving results: {str(e)}")
            else:
                self.status_label.config(text="File not saved. Please select a valid path.")
        else:
            self.status_label.config(text="No analysis results to save.")

# def process_excel_file(file_path):
#     df = pd.read_excel(file_path, sheet_name=0)
#     bases = []
#     for _, row in df.iterrows():
#         base_number = str(row["BSNum"]).strip()
#         bases.append({
#             "Base Name": row["BSName"],
#             "Base Number": base_number,
#             "Latitude": row["Lat"],
#             "Longitude": row["Lon"],
#             "Azimuth 900": row.get("GSM-900", None),
#             "Azimuth 1800": row.get("GSM-1800", None)
#         })
#     return bases

def process_excel_file(file_path):
    df = pd.read_excel(file_path, sheet_name=0)
    bases = []
    for _, row in df.iterrows():
        bases.append({
            "Base Name": row.iloc[1],  # 2-устун
            "Base Number": str(row.iloc[2]).strip(),  # 2-устун
            "Latitude": row.iloc[4],  # 3-устун
            "Longitude": row.iloc[5],  # 5-устун
            "Azimuth 900": row.iloc[6] if len(row) > 4 else None,  # 6-устун
            "Azimuth 1800": row.iloc[7] if len(row) > 5 else None  # 7-устун
        })
    return bases

def sort_conflicts_by_distance(conflicts):
    return sorted(conflicts, key=lambda x: float(x.split('Distance: ')[-1].replace('m', '').strip()))

def analyze_same_frequencies(text_data, excel_data, distance_limit, status_label):
    results = []
    total_bases = len(text_data)
    for i, base_text in enumerate(text_data):
        matched = False
        for base_excel in excel_data:
            if base_text["Base Number"] == base_excel["Base Number"]:
                azimuth_key = f"Azimuth {900 if base_text['Frequency'] < 600 else 1800}"
                azimuth = base_excel.get(azimuth_key, None)
                if azimuth is not None:
                    matched = True
                    conflict = check_frequency_conflict(base_text, text_data, excel_data, distance_limit)

                    if conflict:
                        conflict_sorted = sort_conflicts_by_distance(conflict)
                    else:
                        conflict_sorted = []

                    results.append({
                        "Base Name": base_excel["Base Name"],
                        "Sector": base_text["Sector"],
                        "Frequency": base_text["Frequency"],
                        "Frequency Type": base_text["Frequency Type"],
                        "Azimuth": azimuth,
                        "Frequency Conflict": conflict_sorted
                    })
        if not matched:
            print(f"No match for Base Number: {base_text['Base Number']} in Excel data")
        status_label.config(text=f"Обработка: {i+1}/{total_bases} сектора")
        status_label.update_idletasks()
    return results

def analyze_adjacent_frequencies(text_data, excel_data, distance_limit, status_label):
    results = []
    total_bases = len(text_data)
    for i, base_text in enumerate(text_data):
        matched = False
        for base_excel in excel_data:
            if base_text["Base Number"] == base_excel["Base Number"]:
                azimuth_key = f"Azimuth {900 if base_text['Frequency'] < 600 else 1800}"
                azimuth = base_excel.get(azimuth_key, None)
                if azimuth is not None:
                    matched = True
                    conflict = check_adjacent_frequency_conflict(base_text, text_data, excel_data, distance_limit)

                    if conflict:
                        conflict_sorted = sort_conflicts_by_distance(conflict)
                    else:
                        conflict_sorted = []

                    results.append({
                        "Base Name": base_excel["Base Name"],
                        "Sector": base_text["Sector"],
                        "Frequency": base_text["Frequency"],
                        "Frequency Type": base_text["Frequency Type"],
                        "Azimuth": azimuth,
                        "Frequency Conflict": conflict_sorted
                    })
        if not matched:
            print(f"No match for Base Number: {base_text['Base Number']} in Excel data")
        status_label.config(text=f"Обработка: {i+1}/{total_bases} сектора")
        status_label.update_idletasks()
    return results

def check_frequency_conflict(base_text, text_data, excel_data, distance_limit):
    conflicts = []
    base_coords = next(({
        "Latitude": base["Latitude"],
        "Longitude": base["Longitude"]
    } for base in excel_data if base["Base Number"] == base_text["Base Number"]), None)

    if not base_coords:
        return conflicts

    for other_base in text_data:
        if base_text != other_base and base_text["Frequency"] == other_base["Frequency"]:
            other_coords = next(({
                "Latitude": base["Latitude"],
                "Longitude": base["Longitude"]
            } for base in excel_data if base["Base Number"] == other_base["Base Number"]), None)

            if other_coords:
                distance = geodesic(
                    (base_coords["Latitude"], base_coords["Longitude"]),
                    (other_coords["Latitude"], other_coords["Longitude"])
                ).meters

                if distance <= distance_limit:
                    conflict_str = f"Base: {other_base['Base Number']}, Sector: {other_base['Sector']}, Frequency: {other_base['Frequency']}, Distance: {distance:.2f}m"
                    conflicts.append(conflict_str)

    return conflicts

def check_adjacent_frequency_conflict(base_text, text_data, excel_data, distance_limit):
    conflicts = []
    base_coords = next(({
        "Latitude": base["Latitude"],
        "Longitude": base["Longitude"]
    } for base in excel_data if base["Base Number"] == base_text["Base Number"]), None)

    if not base_coords:
        return conflicts

    for other_base in text_data:
        if base_text != other_base and abs(base_text["Frequency"] - other_base["Frequency"]) == 1:
            other_coords = next(({
                "Latitude": base["Latitude"],
                "Longitude": base["Longitude"]
            } for base in excel_data if base["Base Number"] == other_base["Base Number"]), None)

            if other_coords:
                distance = geodesic(
                    (base_coords["Latitude"], base_coords["Longitude"]),
                    (other_coords["Latitude"], other_coords["Longitude"])
                ).meters

                if distance <= distance_limit:
                    conflict_str = f"Base: {other_base['Base Number']}, Sector: {other_base['Sector']}, Frequency: {other_base['Frequency']}, Distance: {distance:.2f}m"
                    conflicts.append(conflict_str)

    return conflicts

def save_results(results, output_path):
    data = []
    for result in results:
        conflict_data = result.get("Frequency Conflict", [])

        conflict_columns = {}
        for idx, conflict in enumerate(conflict_data):
            conflict_columns[f"Conflict {idx+1}"] = str(conflict)

        row_data = {
            "Base Name": result["Base Name"],
            "Sector": result["Sector"],
            "Frequency": result["Frequency"],
            "Frequency Type": result["Frequency Type"],
            "Azimuth": result["Azimuth"],
        }

        row_data.update(conflict_columns)
        data.append(row_data)

    df = pd.DataFrame(data)

    excel_output_path = output_path.replace(".txt", ".xlsx")

    df.to_excel(excel_output_path, index=False, engine='openpyxl')

    wb = load_workbook(excel_output_path)
    ws = wb.active
    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=ws.max_column):
        for cell in row:
            if 'Distance: 0.00' in str(cell.value):
                cell.fill = fill

    wb.save(excel_output_path)
    print(f"Results saved to {excel_output_path}")
    os.startfile(excel_output_path)

def main():
    global root
    root = Tk()
    root.title("Анализатор Частот")

    welcome_label = Label(root, text="Пожалуйста, выберите", font=("Arial", 14))
    welcome_label.pack(pady=20)

    testsa_button = Button(
        root,
        text="Определение одинаковых частот в заданном радиусе",
        command=open_testsa_window,
        width=55,
        height=2
    )
    testsa_button.pack(pady=10)

    testne_button = Button(
        root,
        text="Определение смежных частот в заданном радиусе",
        command=open_testne_window,
        width=55,
        height=2
    )
    testne_button.pack(pady=10)

    about_button = Button(root, text="О программе", command=show_about, font=("Arial", 9), width=15, height=1)
    about_button.pack(side="left", padx=10, pady=20)

    exit_button = Button(root, text="Выход", command=root.quit, font=("Arial", 9), width=15, height=1)
    exit_button.pack(side="right", padx=10, pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()